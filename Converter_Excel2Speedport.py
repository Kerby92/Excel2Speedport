import openpyxl
import os

os.system("cls")

book = openpyxl.load_workbook('Speedport_Telefonbuch.xlsx')

sheet = book.active

f = open("Speedport_Telefonbuch.txt","w")

counter = 0

for row in sheet.iter_rows(min_row=1, max_col=10, max_row=sheet.max_row):

    for cell in row:
        counter += 1

        if str(cell.value) == "None":
            f.write("\"\"")
        else:
            f.write("\""+str(cell.value) + "\"")

        if counter < 10:
            f.write(",")
        elif counter == 10:
            f.write("\n")
            counter = 0

f.close()

print("Excel-Liste wurde erfolgreich zu Speedport_Telefonbuch.txt konvertiert!\n")
os.system("pause")