# BUGS:
#---------------------------
# offener Bug 23012021/001: leere Zeilen in Excel-Vorlage werden übernommen -> sollten ignoriert werden
#
# offener Bug 23012021/002: Vorlage_Telefonbuch_Bug.xlsx hat scheinbar am Ende der Liste zwei beschriebene Reihen.
#                           Dort steht aber nichts und wird trotzdem übernommen -> Problem untersuchen (Variable sheet.max_row verursacht Probleme)
#
#
# Potentielle Features:
#---------------------------
# Feature 23012021/001: Programm sollte eine Warnung ausgeben, wenn kein Name und Vorname oder keine Rufnummer angegeben wurde
#
# Feature 23012021/002: Kontakte werden im Speedport nicht aktualisiert, falls sich eine Angabe ändert, sondern komplett neu hinzugefügt.
#                       Methode erforderlich, die die Kontaktliste des Speedports vorher leert, um Redudanzen zu vermeiden.
#                       -> Versuche wurden mit Upload2Speedport und "Selenium" durchgeführt, jedoch funktionierte die gebaute *.exe nicht
#                       an einem anderen Windows 7 Computer ohne vollständige Installation mit Python+Selenium


# Import all required functions from the corresponding packes:
#----------------------------------------------------------------------------------------
from openpyxl import load_workbook
from os import system

system("cls") # Clear the command window, if not already empty


# Initialize access variables to the *.xlsx file:
#----------------------------------------------------------------------------------------
try:
    book = load_workbook('Vorlage_Telefonbuch.xlsx')
    
except Exception as e:
    print(e); print("")
    print("Bitte drücken Sie eine beliebige Taste, um die Anwendung zu schließen...")
    system("pause >nul")

sheet = book.active

numContacts = sheet.max_row-1


# Copy and write the data from the *.xlsx to a *.txt file for any cell of a row:
#----------------------------------------------------------------------------------------
if numContacts >= 1:
    f = open("Telefonbuch.txt", mode='w', encoding='utf-8')

    for row in sheet.iter_rows(min_row=1, max_col=10, max_row=sheet.max_row):
        
        counter = 0
        bitMask = 0
        for cell in row:
            if str(cell.value) == "None":
                bitMask = bitMask | 0x200 >> counter
            counter += 1

        # Check, if a row is totally empty:
        if bitMask == 0x3FF: # print(format(0x3FF,'#012b'))
            numContacts -= 1

            # If an empty row leads to zero contacts in the list, stop the for-loop:
            if numContacts == 0:
                break

            # Otherwise just skip this row-loop, since this row is empty:
            continue

        # Check, if contact phone numbers were correctly supplied:
        if bitMask & 0xF0 == 0xF0:
            print(">>> Warnung! Es wurde keine Telefonnummer für den Kontakt in Zeile " + str(cell.row) + " hinterlegt. Der Kontakt wird übersprungen! <<<\n")
            numContacts -= 1
            continue

        # Check, if contact name was correctly supplied:
        if bitMask & 0x300 == 0x300:
            print(">>> Warnung! Es wurde kein Name für den Kontakt in Zeile " + str(cell.row) + " hinterlegt. Der Kontakt wird übersprungen! <<<\n")
            numContacts -= 1
            continue

        counter = 0
        for cell in row:
            counter += 1

            if str(cell.value) == "None":
                f.write("\"\"")

            else:
                f.write("\""+str(cell.value) + "\"")

            if counter < 10:
                f.write(",")
            elif counter == 10:
                f.write("\n")
                counter = 0

f.close()


# Check the number of supplied contacts in the *.xlsx and depending on the number,
# generate a txt or tell user, that there was no contact supplied.
#----------------------------------------------------------------------------------------
if numContacts > 1:
    print("Es wurden insgesamt " + str(numContacts) + " Kontakte erkannt.\n")

elif numContacts == 0:
    print("Die Excel-Liste \"Vorlage_Telefonbuch.xlsx\" beinhaltet keine Kontakte.\n")

    print("Bitte drücken Sie eine beliebige Taste, um die Anwendung zu schließen...")
    system("pause >nul")
    exit()

else:
    print("Es wurde " + str(numContacts) + " Kontakt erkannt.\n")


# Print out the finish messages:
#----------------------------------------------------------------------------------------
print("Excel-Liste wurde erfolgreich zu \"Telefonbuch.txt\" im UTF-8 Format konvertiert!")

print("Sie können \"Telefonbuch.txt\" über die Benutzeroberfläche des Speedports \"www.speedport.ip\" importieren. \n")

print("Bitte drücken Sie eine beliebige Taste, um die Anwendung zu schließen...")

system("pause >nul") # ">nul" suppresses the default message by the system based pause function